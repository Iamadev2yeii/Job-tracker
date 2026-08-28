"""
Manages the persistent Excel tracker file. Four sheets:

  1. "Jobs" — sustainability/ESG-relevant roles in Munich, picked
     directly from company career pages (no job boards/agencies) plus
     the Bundesagentur API (config/cv_profile.yaml).
  2. "Munich Internships & Trainee" — Praktikum/Trainee
     postings in Munich, any field (not just sustainability, not just
     office/admin — deliberately broad on field, narrow on seniority).
     Werkstudent roles are excluded. Still scored for how well it
     touches Prabha's actual CV themes, purely to rank, never to
     exclude (config/general_roles_profile.yaml). This sheet was
     previously named "General Roles" and covered generic office/
     admin work at any seniority — if your existing tracker file still
     has that old sheet, it's now an inert leftover (no longer
     updated); delete it by hand in Excel if you want it gone, same as
     any other renamed-sheet leftover in this project.
  3. "Remote Sustainability (Europe)" — fully remote ESG/sustainability
     roles anywhere in Europe, not scoped to Munich
     (config/remote_sustainability_profile.yaml).
  4. "Jobs - Berlin" — sustainability/ESG-relevant roles in Berlin,
     same philosophy as sheet 1 but for Berlin's metro area instead of
     Munich's (config/berlin_companies.yaml,
     config/berlin_arbeitsagentur_searches.yaml). Added because
     Berlin reportedly has a much larger sustainability/climate-tech
     company base than Munich.

No relevance floor is ever applied on sheets 2 or 3 by default — the
point of those sheets is employability/reach, not specialty fit.

All four sheets share the same behaviour:
  - Each run, only genuinely NEW postings (by URL) are added.
  - Postings already in a sheet are left alone (deduped by URL, per
    sheet — a URL only needs to be unique within its own sheet).
  - Every run, rows scoring below that sheet's min_score are pruned —
    applies to rows already sitting in the sheet too. All four
    default to 0 (off).
  - Every run, rows whose title now fails the universal seniority or
    language exclusion checks (src/matcher.py) are also pruned — a
    rule tightened after a bad row was already added (e.g. a language
    exclusion added after a Korean-language posting made it in) would
    otherwise leave that row sitting there forever, since new runs
    only ever add rows, they don't re-check what's already there.
  - Each sheet is re-sorted by Relevance Score (highest first) every
    run, so the best matches are always at the top.
  - archive_and_reset() (called from reset_tracker.py) archives ALL
    FOUR sheets together as one file and starts fresh for all of them.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

COLUMNS = [
    "Job Posted",
    "Company",
    "Job Title",
    "Relevance Score (1-10)",
    "Location",
    "URL",
]
COLUMN_WIDTHS = [14, 26, 44, 12, 30, 60]

# Old header names that should map onto a current column, so a rename
# doesn't strand existing data.
HEADER_ALIASES: dict[str, str] = {}

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NEW_ROW_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # green
GENERAL_NEW_ROW_FILL = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")  # amber
REMOTE_NEW_ROW_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # blue
BERLIN_NEW_ROW_FILL = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")  # light purple

SHEET_NAME = "Jobs"
GENERAL_SHEET_NAME = "Munich Internships & Trainee"
REMOTE_SHEET_NAME = "Remote Sustainability (Europe)"
BERLIN_SHEET_NAME = "Jobs - Berlin"

ALL_SHEETS = [SHEET_NAME, GENERAL_SHEET_NAME, REMOTE_SHEET_NAME, BERLIN_SHEET_NAME]


def _style_header(ws: Worksheet) -> None:
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    for i, w in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _new_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(COLUMNS)
    _style_header(ws)

    for name in ALL_SHEETS[1:]:
        sheet = wb.create_sheet(name)
        sheet.append(COLUMNS)
        _style_header(sheet)
    return wb


def _migrate_sheet(wb: Workbook, sheet_name: str, sheet_index: int) -> None:
    ws = wb[sheet_name]
    existing_headers = [c.value for c in ws[1]]
    if existing_headers == COLUMNS:
        return  # already current

    mapped_headers = [HEADER_ALIASES.get(h, h) for h in existing_headers]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_dict = {}
        for col_idx, header in enumerate(mapped_headers, start=1):
            if header in COLUMNS:
                row_dict[header] = ws.cell(row=row_idx, column=col_idx).value
        if row_dict:
            rows.append(row_dict)

    wb.remove(ws)
    new_ws = wb.create_sheet(sheet_name, sheet_index)
    new_ws.append(COLUMNS)
    for row_dict in rows:
        new_ws.append([row_dict.get(col, "") for col in COLUMNS])
    _style_header(new_ws)


def _ensure_sheet(wb: Workbook, sheet_name: str, sheet_index: int) -> None:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name, sheet_index)
        ws.append(COLUMNS)
        _style_header(ws)
    else:
        _migrate_sheet(wb, sheet_name, sheet_index)


def load_or_create(path: Path) -> Workbook:
    if not path.exists():
        return _new_workbook()
    wb = load_workbook(path)
    for i, name in enumerate(ALL_SHEETS):
        _ensure_sheet(wb, name, i)
    return wb


def _existing_urls(ws: Worksheet) -> set[str]:
    url_col = COLUMNS.index("URL") + 1
    return {
        ws.cell(row=row, column=url_col).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=url_col).value
    }


def _prune_below_score(ws: Worksheet, min_score: int) -> int:
    """Deletes any row scoring below min_score. Runs every update, so
    it retroactively cleans rows already in the sheet too. Returns how
    many rows were removed."""
    if not min_score:
        return 0
    score_col = COLUMNS.index("Relevance Score (1-10)") + 1
    rows_to_delete = [
        row
        for row in range(2, ws.max_row + 1)
        if not (
            isinstance(ws.cell(row=row, column=score_col).value, (int, float))
            and ws.cell(row=row, column=score_col).value >= min_score
        )
    ]
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
    return len(rows_to_delete)


def _prune_excluded_titles(ws: Worksheet, revalidate_fn: "Callable[[str, str], bool] | None" = None) -> int:
    """
    Deletes any row whose Job Title now fails the universal seniority
    or language exclusion checks (src/matcher.py), OR (if revalidate_fn
    is given) no longer passes its sheet's own title-relevance check.

    The revalidate_fn addition matters a lot in practice: a real run
    let the Bundesagentur searches skip title-relevance checking
    entirely for a while, which flooded every sheet with hundreds of
    industry-adjacent-but-irrelevant rows (electricians, warehouse
    staff, etc. from broad searches like "Windenergie"). Fixing the
    matching logic only stops NEW bad rows — without this, all of
    that existing noise would sit in the sheet forever, since a run
    only ever adds rows, it doesn't normally re-check what's already
    there. revalidate_fn is built in main.py from each sheet's actual
    profile, so this cleans up past runs' mistakes too, not just
    prevents future ones.

    Only the Job Title and Company columns are available for this
    re-check (the tracker doesn't store description text) — a row
    that was originally matched via description text rather than its
    own title (see src/matcher.py, _combined_text) could theoretically
    get pruned here even though it was a legitimate match. That's an
    accepted, conservative trade-off given how much worse the
    alternative (leaving hundreds of confirmed-irrelevant rows in
    place) is.
    """
    from src.matcher import is_excluded_seniority, is_excluded_language

    title_col = COLUMNS.index("Job Title") + 1
    company_col = COLUMNS.index("Company") + 1
    rows_to_delete = []
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=title_col).value
        if not title:
            continue
        if is_excluded_seniority(title) or is_excluded_language(title):
            rows_to_delete.append(row)
            continue
        if revalidate_fn is not None:
            company = ws.cell(row=row, column=company_col).value
            if not revalidate_fn(title, company):
                rows_to_delete.append(row)
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
    return len(rows_to_delete)


def _sort_by_relevance(ws: Worksheet, newly_added_urls: set[str], fill: PatternFill) -> None:
    """Re-sorts all data rows by Relevance Score (highest first), then
    re-applies the "new" highlight to whichever URLs were added this
    run (their row position may have moved during the sort)."""
    score_col = COLUMNS.index("Relevance Score (1-10)")
    url_col = COLUMNS.index("URL")

    rows = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, len(COLUMNS) + 1)]
        rows.append(values)

    rows.sort(key=lambda v: (v[score_col] if isinstance(v[score_col], (int, float)) else 0), reverse=True)

    for row in range(2, ws.max_row + 1):
        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=row, column=col).value = None
            ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)

    for i, values in enumerate(rows):
        row_idx = i + 2
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value
        if values[url_col] in newly_added_urls:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill


def _build_row(job: dict[str, Any], url: str) -> list:
    return [
        job.get("posted_date", ""),
        job.get("company", ""),
        job.get("title", ""),
        job.get("relevance_score", 1),
        job.get("location", ""),
        url,
    ]


def _update_sheet(
    wb: Workbook, sheet_name: str, fill: PatternFill, new_jobs: list[dict[str, Any]], min_score: int,
    revalidate_fn: "Callable[[str, str], bool] | None" = None,
) -> dict[str, int]:
    ws = wb[sheet_name]
    existing_urls = _existing_urls(ws)

    added = 0
    already_tracked = 0
    newly_added_urls = set()
    for job in new_jobs:
        url = job.get("url", "")
        if not url:
            continue
        if url in existing_urls:
            already_tracked += 1
            continue
        ws.append(_build_row(job, url))
        newly_added_urls.add(url)
        existing_urls.add(url)
        added += 1

    pruned = _prune_below_score(ws, min_score)
    pruned_excluded = _prune_excluded_titles(ws, revalidate_fn)
    _sort_by_relevance(ws, newly_added_urls, fill)
    return {
        "added": added, "already_tracked": already_tracked,
        "pruned": pruned + pruned_excluded, "total_rows": ws.max_row - 1,
    }


def update_tracker(
    path: Path, new_jobs: list[dict[str, Any]], min_score: int = 0,
    revalidate_fn: "Callable[[str, str], bool] | None" = None,
) -> dict[str, int]:
    """Updates the "Jobs" (Munich sustainability) sheet. Saves the file."""
    wb = load_or_create(path)
    summary = _update_sheet(wb, SHEET_NAME, NEW_ROW_FILL, new_jobs, min_score, revalidate_fn)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return summary


def update_general_tracker(
    path: Path, new_jobs: list[dict[str, Any]], min_score: int = 0,
    revalidate_fn: "Callable[[str, str], bool] | None" = None,
) -> dict[str, int]:
    """Updates the "Munich Internships & Trainee" sheet. Saves
    the file. Called separately from update_tracker — each opens,
    updates its own sheet, and saves; the file just gets saved
    multiple times per run, which is harmless."""
    wb = load_or_create(path)
    summary = _update_sheet(wb, GENERAL_SHEET_NAME, GENERAL_NEW_ROW_FILL, new_jobs, min_score, revalidate_fn)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return summary


def update_remote_tracker(
    path: Path, new_jobs: list[dict[str, Any]], min_score: int = 0,
    revalidate_fn: "Callable[[str, str], bool] | None" = None,
) -> dict[str, int]:
    """Updates the "Remote Sustainability (Europe)" sheet. Saves the
    file. Called separately from the other two update_* functions."""
    wb = load_or_create(path)
    summary = _update_sheet(wb, REMOTE_SHEET_NAME, REMOTE_NEW_ROW_FILL, new_jobs, min_score, revalidate_fn)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return summary


def update_berlin_tracker(
    path: Path, new_jobs: list[dict[str, Any]], min_score: int = 0,
    revalidate_fn: "Callable[[str, str], bool] | None" = None,
) -> dict[str, int]:
    """Updates the "Jobs - Berlin" sheet. Saves the file. Called
    separately from the other update_* functions."""
    wb = load_or_create(path)
    summary = _update_sheet(wb, BERLIN_SHEET_NAME, BERLIN_NEW_ROW_FILL, new_jobs, min_score, revalidate_fn)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return summary


def archive_and_reset(path: Path, archive_dir: Path) -> Path | None:
    """
    Moves the current tracker (all four sheets) to
    archive/job_tracker_YYYY-MM.xlsx and creates a fresh empty
    tracker (all four sheets, empty) at `path`. Returns the archive
    path, or None if there was nothing to archive.
    """
    if not path.exists():
        _new_workbook().save(path)
        return None

    archive_dir.mkdir(parents=True, exist_ok=True)
    month_tag = dt.date.today().strftime("%Y-%m")
    archive_path = archive_dir / f"job_tracker_{month_tag}.xlsx"

    counter = 2
    final_path = archive_path
    while final_path.exists():
        final_path = archive_dir / f"job_tracker_{month_tag}_v{counter}.xlsx"
        counter += 1

    path.rename(final_path)
    _new_workbook().save(path)
    return final_path
